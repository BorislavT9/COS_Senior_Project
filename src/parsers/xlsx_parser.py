"""
XLSX parser: extract cell values from first sheet as text (row-by-row).
"""

from pathlib import Path

from openpyxl import load_workbook


def parse_xlsx(file_path: str) -> str:
    """
    Extract cell values from first sheet of XLSX at file_path.
    Returns a text representation: one row per line, cells separated by tab.
    """
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"XLSX not found: {file_path}")
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheet = wb.active
    if sheet is None:
        wb.close()
        return ""
    rows = []
    for row in sheet.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        rows.append("\t".join(cells))
    wb.close()
    return "\n".join(rows)
